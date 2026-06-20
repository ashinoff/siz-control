"""Report building and Excel/CSV export."""
import csv
import io
from datetime import date
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from ..enums import (
    DeadlineStatus,
    InventoryStatus,
    VerificationStatus,
)
from ..models.inventory import InventoryItem
from . import status as status_service

# Human-readable labels --------------------------------------------------

TYPE_LABELS = {"ppe": "СИЗ", "material": "Материал", "equipment": "Оборудование"}
STATUS_LABELS = {
    "in_stock": "На складе",
    "issued": "У сотрудника",
    "to_writeoff": "К списанию",
    "written_off": "Списано",
}
DEADLINE_LABELS = {
    "in_date": "В сроке",
    "expiring": "Скоро истекает",
    "expired": "Просрочено",
    "not_started": "Не начата",
    "not_applicable": "—",
}
VERIFICATION_LABELS = {
    "in_date": "В сроке",
    "expiring": "Скоро истекает",
    "expired": "Просрочено",
    "not_required": "Не требуется",
}


def _row(item: InventoryItem, today: date) -> Dict[str, object]:
    catalog = item.catalog_item
    return {
        "ID": item.id,
        "Тип": TYPE_LABELS.get(item.item_type, item.item_type),
        "Категория": catalog.category.name if catalog and catalog.category else "",
        "Наименование": catalog.name if catalog else "",
        "Инв. номер": item.inventory_number or "",
        "Серийный номер": item.serial_number or "",
        "Кол-во": item.quantity,
        "Подразделение": item.department_owner.name if item.department_owner else "",
        "Склад": item.current_warehouse.name if item.current_warehouse else "",
        "Сотрудник": item.current_employee.full_name if item.current_employee else "",
        "Статус": STATUS_LABELS.get(item.status, item.status),
        "Начало экспл.": item.service_start_date.isoformat() if item.service_start_date else "",
        "Окончание экспл.": item.service_end_date.isoformat() if item.service_end_date else "",
        "Статус срока": DEADLINE_LABELS.get(status_service.deadline_status(item, today).value, ""),
        "Поверка до": item.next_verification_date.isoformat() if item.next_verification_date else "",
        "Статус поверки": VERIFICATION_LABELS.get(
            status_service.verification_status(item, today).value, ""
        ),
    }


def _scoped_query(db: Session, scope_department_id: Optional[int]):
    query = (
        db.query(InventoryItem)
        .options(
            joinedload(InventoryItem.catalog_item),
            joinedload(InventoryItem.department_owner),
            joinedload(InventoryItem.current_warehouse),
            joinedload(InventoryItem.current_employee),
        )
        .filter(InventoryItem.is_active.is_(True))
    )
    if scope_department_id is not None:
        query = query.filter(InventoryItem.department_owner_id == scope_department_id)
    return query


REPORTS = {
    "stock": "Остатки на складах",
    "issued": "Выдано сотрудникам",
    "expired": "Просроченные по эксплуатации",
    "expiring": "Истекающие сроки (30 дней)",
    "verification_expired": "Просроченная поверка",
    "verification_expiring": "Поверка в ближайшие 30 дней",
    "all": "Все позиции",
}


def build_report(
    db: Session,
    report: str,
    *,
    scope_department_id: Optional[int] = None,
    department_id: Optional[int] = None,
    employee_id: Optional[int] = None,
) -> List[Dict[str, object]]:
    today = date.today()
    query = _scoped_query(db, scope_department_id)
    if department_id is not None and scope_department_id is None:
        query = query.filter(InventoryItem.department_owner_id == department_id)
    if employee_id is not None:
        query = query.filter(InventoryItem.current_employee_id == employee_id)

    items = query.all()
    rows: List[Dict[str, object]] = []
    for item in items:
        d = status_service.deadline_status(item, today)
        v = status_service.verification_status(item, today)
        if report == "stock" and item.status != InventoryStatus.IN_STOCK.value:
            continue
        if report == "issued" and item.status != InventoryStatus.ISSUED.value:
            continue
        if report == "expired" and d != DeadlineStatus.EXPIRED:
            continue
        if report == "expiring" and d != DeadlineStatus.EXPIRING:
            continue
        if report == "verification_expired" and v != VerificationStatus.EXPIRED:
            continue
        if report == "verification_expiring" and v != VerificationStatus.EXPIRING:
            continue
        rows.append(_row(item, today))
    return rows


def to_csv(rows: List[Dict[str, object]]) -> bytes:
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    # BOM so Excel opens UTF-8 correctly.
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def to_xlsx(rows: List[Dict[str, object]], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Отчет"

    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row[h] for h in headers])
        # Auto width.
        for col, header in enumerate(headers, start=1):
            max_len = max([len(str(header))] + [len(str(r[header])) for r in rows])
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)
        ws.freeze_panes = "A2"
    else:
        ws.append(["Нет данных"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
