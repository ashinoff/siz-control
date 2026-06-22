"""Import inventory items from Excel (.xlsx) files — admin only."""
import io
from datetime import date
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import require_admin
from ..enums import InventoryStatus, LifeUnit
from ..models.catalog import CatalogItem, Category
from ..models.inventory import InventoryItem
from ..models.organization import Warehouse
from ..models.user import User
from ..services import status as status_service
from ..services.audit import log_movement
from ..enums import OperationType

router = APIRouter(prefix="/api/import", tags=["import"])

# Column mapping: Excel header → field handling
TEMPLATE_COLUMNS = [
    "Тип (ppe/material/equipment)",
    "Категория",
    "Наименование",
    "Инв. номер",
    "Серийный номер",
    "Кол-во",
    "Подразделение (код склада ID)",
    "Склад (ID)",
    "Дата поступления (ГГГГ-ММ-ДД)",
    "Срок службы (число)",
    "Ед. срока (days/months/years)",
    "Требует поверки (да/нет)",
    "Комментарий",
]


@router.get("/template")
def download_template(_: User = Depends(require_admin)):
    """Download an Excel template for import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт"

    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = max(len(name) + 2, 15)

    # Example row
    example = ["ppe", "Перчатки", "Перчатки диэлектрические", "ИНВ-001", "СН-12345",
               2, 1, 1, "2025-01-15", 12, "months", "нет", "Пример"]
    ws.append(example)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="import_template.xlsx"'},
    )


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val).strip()[:10])
    except (ValueError, TypeError):
        return None


@router.post("/upload")
def upload_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(require_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Допустимы только файлы .xlsx")

    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать файл Excel")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Файл пуст (нет строк данных)")

    # Cache lookups
    categories_cache = {}
    catalog_cache = {}

    created = 0
    errors: List[str] = []

    for i, row in enumerate(rows, start=2):
        if len(row) < 6:
            errors.append(f"Строка {i}: недостаточно столбцов")
            continue

        item_type = str(row[0] or "").strip().lower()
        if item_type not in ("ppe", "material", "equipment"):
            errors.append(f"Строка {i}: неверный тип '{row[0]}' (допустимо: ppe, material, equipment)")
            continue

        category_name = str(row[1] or "").strip()
        name = str(row[2] or "").strip()
        if not name:
            errors.append(f"Строка {i}: наименование пусто")
            continue

        inv_number = str(row[3] or "").strip() or None
        serial_number = str(row[4] or "").strip() or None
        try:
            quantity = int(row[5]) if row[5] is not None else 1
        except (ValueError, TypeError):
            quantity = 1

        try:
            dept_id = int(row[6]) if row[6] is not None else None
        except (ValueError, TypeError):
            errors.append(f"Строка {i}: неверный ID подразделения '{row[6]}'")
            continue

        try:
            wh_id = int(row[7]) if row[7] is not None else None
        except (ValueError, TypeError):
            wh_id = None

        date_received = _parse_date(row[8]) if len(row) > 8 else None

        try:
            life_value = int(row[9]) if len(row) > 9 and row[9] is not None else None
        except (ValueError, TypeError):
            life_value = None

        life_unit_raw = str(row[10] or "").strip().lower() if len(row) > 10 else None
        life_unit = life_unit_raw if life_unit_raw in ("days", "months", "years") else None

        req_verif_raw = str(row[11] or "").strip().lower() if len(row) > 11 else "нет"
        requires_verification = req_verif_raw in ("да", "yes", "true", "1")

        comment = str(row[12] or "").strip() or None if len(row) > 12 else None

        if dept_id is None:
            errors.append(f"Строка {i}: не указано подразделение")
            continue

        # Find or create catalog item
        cache_key = (item_type, category_name, name)
        if cache_key in catalog_cache:
            catalog_item = catalog_cache[cache_key]
        else:
            # Find category
            category = None
            if category_name:
                if category_name not in categories_cache:
                    cat = db.query(Category).filter(
                        Category.name == category_name,
                        Category.item_type == item_type,
                        Category.is_active.is_(True),
                    ).first()
                    if not cat:
                        cat = Category(name=category_name, item_type=item_type)
                        db.add(cat)
                        db.flush()
                    categories_cache[category_name] = cat
                category = categories_cache[category_name]

            # Find or create catalog item
            q = db.query(CatalogItem).filter(
                CatalogItem.name == name,
                CatalogItem.item_type == item_type,
                CatalogItem.is_active.is_(True),
            )
            if category:
                q = q.filter(CatalogItem.category_id == category.id)
            catalog_item = q.first()
            if not catalog_item:
                catalog_item = CatalogItem(
                    name=name,
                    item_type=item_type,
                    category_id=category.id if category else None,
                    life_value=life_value,
                    life_unit=life_unit or LifeUnit.MONTHS.value,
                    requires_verification=requires_verification,
                )
                db.add(catalog_item)
                db.flush()
            catalog_cache[cache_key] = catalog_item

        # Create inventory item
        inv = InventoryItem(
            catalog_item_id=catalog_item.id,
            item_type=item_type,
            inventory_number=inv_number,
            serial_number=serial_number,
            quantity=quantity,
            department_owner_id=dept_id,
            current_warehouse_id=wh_id,
            status=InventoryStatus.IN_STOCK.value,
            date_received=date_received,
            life_value=life_value,
            life_unit=life_unit,
            requires_verification=requires_verification,
            comment=comment,
        )
        if inv.life_starts_in_stock and date_received:
            inv.service_start_date = date_received
            status_service.recalc_service_dates(inv)

        db.add(inv)
        db.flush()

        log_movement(
            db,
            user_id=current.id,
            operation_type=OperationType.CREATE.value,
            inventory_item_id=inv.id,
            department_id=dept_id,
            object_label=name,
            comment="Импорт из Excel",
        )
        created += 1

    db.commit()

    result = {"detail": f"Импортировано: {created} позиций", "created": created}
    if errors:
        result["errors"] = errors
        result["detail"] += f", ошибок: {len(errors)}"
    return result
