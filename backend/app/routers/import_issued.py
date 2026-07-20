"""Import issued equipment registers from Excel — admin only.

Validates:
- Employee FIO exists in DB
- Catalog item (nomenclature) exists
- Employee belongs to the specified department
If all pass → creates inventory item with status ISSUED.
If not → collects errors and returns error report.
"""
import io
import json
from datetime import date
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import require_admin
from ..enums import InventoryStatus, OperationType
from ..models.catalog import CatalogItem
from ..models.inventory import InventoryItem
from ..models.organization import Department, Employee, Warehouse
from ..models.user import User
from ..services.audit import log_movement

router = APIRouter(prefix="/api/import-issued", tags=["import"])

TEMPLATE_COLUMNS = [
    "ФИО сотрудника",
    "Подразделение (РЭС)",
    "Наименование СИЗ/СИ",
    "Марка / тип",
    "Инв. номер",
    "Серийный номер",
    "Год выпуска",
    "Класс точности",
    "Предел (диапазон) измерений",
    "Вид КМХ (поверка/калибровка/контроль исправности)",
    "Периодичность КМХ (мес.)",
    "Кол-во",
    "Дата выдачи (ГГГГ-ММ-ДД)",
    "Дата испытания (ГГГГ-ММ-ДД)",
    "Дата след. испытания (ГГГГ-ММ-ДД)",
    "№ свидетельства о поверке",
    "Результат осмотра (годен/негоден/ремонт)",
    "Сведения о ремонтах",
    "Примечание",
]


@router.get("/template")
def download_issued_template(_: User = Depends(require_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр выданного"

    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = max(len(name) + 2, 15)

    example = [
        "Иванов Иван Иванович",
        "Адлерский РЭС",
        "Амперметр Э59",
        "Э59",
        "СИ-100",
        "SN-001",
        2018,
        "0,5",
        "0-600А",
        "поверка",
        12,
        1,
        "2025-06-01",
        "2025-05-15",
        "2025-11-15",
        "№ 1234-2025",
        "годен",
        "",
        "",
    ]
    ws.append(example)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="issued_register_template.xlsx"'},
    )


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val).strip()[:10])
    except (ValueError, TypeError):
        return None


INSPECTION_MAP = {
    "годен": "good",
    "негоден": "failed",
    "ремонт": "repair",
    "требует ремонта": "repair",
    "good": "good",
    "failed": "failed",
    "repair": "repair",
}


@router.post("/upload")
def upload_issued_register(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(require_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Допустимы только файлы .xlsx")

    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать файл Excel")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Файл пуст (нет строк данных)")

    # Pre-load lookups
    all_employees = db.query(Employee).filter(Employee.is_active.is_(True)).all()
    emp_by_name = {}
    for e in all_employees:
        emp_by_name.setdefault(e.full_name.strip().lower(), []).append(e)

    all_departments = {d.name.strip().lower(): d for d in db.query(Department).filter(Department.is_active.is_(True)).all()}

    all_catalog = db.query(CatalogItem).filter(CatalogItem.is_active.is_(True)).all()
    cat_by_name = {}
    for c in all_catalog:
        cat_by_name.setdefault(c.name.strip().lower(), []).append(c)

    created = 0
    skipped = 0
    errors: List[dict] = []

    for i, row in enumerate(rows, start=2):
        if not row or len(row) < 3:
            continue
        # Skip empty rows
        if not any(row[:3]):
            continue

        def cell(idx):
            return row[idx] if len(row) > idx else None

        def cell_str(idx):
            return str(cell(idx) or "").strip()

        def cell_int(idx):
            v = cell(idx)
            try:
                return int(v) if v not in (None, "") else None
            except (ValueError, TypeError):
                return None

        fio = cell_str(0)
        dept_name = cell_str(1)
        item_name = cell_str(2)
        brand_model = cell_str(3)
        inv_number = cell_str(4)
        serial_number = cell_str(5)
        manufacture_year = cell_int(6)
        accuracy_class = cell_str(7)
        measurement_range = cell_str(8)
        metrology_type = cell_str(9)
        metrology_interval_months = cell_int(10)
        quantity = cell_int(11) or 1
        date_issued = _parse_date(cell(12))
        date_test = _parse_date(cell(13))
        date_next_test = _parse_date(cell(14))
        verification_certificate = cell_str(15)
        inspection_raw = cell_str(16).lower()
        inspection_result = INSPECTION_MAP.get(inspection_raw)
        repair_info = cell_str(17)
        comment = cell_str(18)

        row_errors = []

        # Validate FIO
        if not fio:
            row_errors.append("ФИО не указано")
        emp_matches = emp_by_name.get(fio.lower(), [])
        if fio and not emp_matches:
            row_errors.append(f"Сотрудник «{fio}» не найден в базе")

        # Validate department
        dept = all_departments.get(dept_name.lower()) if dept_name else None
        if dept_name and not dept:
            row_errors.append(f"Подразделение «{dept_name}» не найдено")

        # Validate employee in department
        employee = None
        if emp_matches and dept:
            employee = next((e for e in emp_matches if e.department_id == dept.id), None)
            if not employee:
                row_errors.append(f"Сотрудник «{fio}» не числится в «{dept_name}»")
        elif emp_matches and not dept:
            employee = emp_matches[0]

        # Validate nomenclature
        if not item_name:
            row_errors.append("Наименование не указано")
        cat_matches = cat_by_name.get(item_name.lower(), [])
        # Среди одноимённых позиций предпочитаем ту, где задана подкатегория.
        catalog_item = next((c for c in cat_matches if c.subcategory_id), None) \
            or (cat_matches[0] if cat_matches else None)
        if item_name and not catalog_item:
            row_errors.append(f"Номенклатура «{item_name}» не найдена в справочнике")

        if row_errors:
            errors.append({"row": i, "fio": fio, "item": item_name, "errors": row_errors})
            continue

        # Duplicate check: same catalog + inv_number/serial + employee
        dup_q = db.query(InventoryItem).filter(
            InventoryItem.catalog_item_id == catalog_item.id,
            InventoryItem.current_employee_id == employee.id,
            InventoryItem.is_active.is_(True),
        )
        if inv_number:
            dup_q = dup_q.filter(InventoryItem.inventory_number == inv_number)
        if serial_number:
            dup_q = dup_q.filter(InventoryItem.serial_number == serial_number)
        if inv_number or serial_number:
            if dup_q.first():
                skipped += 1
                continue

        # Find warehouse for the department
        wh = db.query(Warehouse).filter(
            Warehouse.department_id == dept.id, Warehouse.is_active.is_(True)
        ).first()

        inv = InventoryItem(
            catalog_item_id=catalog_item.id,
            item_type=catalog_item.item_type,
            inventory_number=inv_number or None,
            serial_number=serial_number or None,
            brand_model=brand_model or None,
            quantity=quantity,
            department_owner_id=dept.id,
            current_warehouse_id=None,
            current_employee_id=employee.id,
            status=InventoryStatus.ISSUED.value,
            date_issued=date_issued,
            date_received=date_issued,
            service_start_date=date_issued,
            manufacture_year=manufacture_year,
            accuracy_class=accuracy_class or None,
            measurement_range=measurement_range or None,
            metrology_type=metrology_type or None,
            metrology_interval_months=metrology_interval_months,
            last_verification_date=date_test,
            next_verification_date=date_next_test,
            verification_certificate=verification_certificate or None,
            last_inspection_result=inspection_result,
            repair_info=repair_info or None,
            requires_verification=catalog_item.requires_verification,
            comment=comment or None,
        )
        db.add(inv)
        db.flush()

        log_movement(
            db,
            user_id=current.id,
            operation_type=OperationType.ISSUE.value,
            inventory_item_id=inv.id,
            department_id=dept.id,
            employee_id=employee.id,
            object_label=catalog_item.name,
            comment="Импорт реестра выданного",
        )
        created += 1

    db.commit()

    result = {
        "detail": f"Импортировано: {created}, дубликатов пропущено: {skipped}",
        "created": created,
        "skipped": skipped,
        "error_count": len(errors),
    }
    if errors:
        result["errors"] = errors
        result["detail"] += f", ошибок: {len(errors)}"
    return result


@router.post("/errors-report")
def download_errors_report(
    errors: List[dict],
    _: User = Depends(require_admin),
):
    """Generate Excel report of import errors."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки импорта"

    header_fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Строка", "ФИО", "Наименование", "Причины ошибки"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 30

    for err in errors:
        ws.append([err["row"], err["fio"], err["item"], "; ".join(err["errors"])])

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="import_errors.xlsx"'},
    )
